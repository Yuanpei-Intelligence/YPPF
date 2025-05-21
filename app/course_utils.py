"""
course_utils.py

course_views.py的依赖函数

registration_status_check: 检查学生选课状态变化的合法性
registration_status_change: 改变学生选课状态
course_to_display: 把课程信息转换为方便前端呈现的形式
draw_lots: 预选阶段结束时执行抽签
change_course_status: 改变课程的选课阶段
remaining_willingness_point（暂不启用）: 计算学生剩余的意愿点数
process_time: 把datetime对象转换成人类可读的时间表示
check_course_time_conflict: 检查当前选择的课是否与已选的课上课时间冲突
"""
from app.utils_dependency import *
from app.models import (
    User,
    NaturalPerson,
    Activity,
    Notification,
    ActivityPhoto,
    Position,
    Participation,
    Course,
    CourseTime,
    CourseParticipant,
    CourseRecord,
    Semester,
)
from app.utils import (
    get_person_or_org,
    if_image,
)
from app.notification_utils import (
    bulk_notification_create,
    notification_create,
    notification_status_change,
)
from app.activity_utils import (
    changeActivityStatus,
    notifyActivity,
    create_participate_infos,
)
from app.extern.wechat import WechatApp, WechatMessageLevel
from app.log import logger

import openpyxl
import openpyxl.worksheet.worksheet
from random import sample
from urllib.parse import quote
from collections import Counter
from datetime import datetime, timedelta
from typing import Tuple, List

from django.http import HttpRequest, HttpResponse, QueryDict
from django.db import IntegrityError, transaction
from django.db.models import F, Sum, Prefetch

from scheduler.adder import ScheduleAdder, MultipleAdder
from scheduler.cancel import remove_job
from utils.config.cast import str_to_time
from achievement.api import unlock_achievement

__all__ = [
    'check_ac_time_course',
    'course_activity_base_check',
    'create_single_course_activity',
    'modify_course_activity',
    'cancel_course_activity',
    'registration_status_change',
    'course_to_display',
    'change_course_status',
    'course_base_check',
    'create_course',
    'cal_participate_num',
    'check_post_and_modify',
    'finish_course',
    'download_course_record',
    'download_select_info',
]

APP_CONFIG = CONFIG.course


def check_ac_time_course(start_time: datetime, end_time: datetime) -> bool:
    """
    时间合法性的检查：开始早于结束

    :param start_time: 活动开始时间
    :type start_time: datetime
    :param end_time: 活动结束时间
    :type end_time: datetime
    :return: 是否合法
    :rtype: bool
    """
    if not start_time < end_time:
        return False
    return True


def course_activity_base_check(request: HttpRequest) -> dict:
    """
    检查课程活动，是activity_base_check的简化版，失败时抛出AssertionError

    :param request: 修改/发起单次课程活动的请求
    :type request: HttpRequest
    :raises AssertionError: 活动时间非法/需要报名的活动必须提前至少一小时发起
    :return: context
    :rtype: dict
    """
    context = dict()

    # 读取活动名称和地点，检查合法性
    context["title"] = request.POST.get("title", "")
    # context["introduction"] = request.POST["introduction"] # 暂定不需要简介
    context["location"] = request.POST.get("location", "")
    assert len(context["title"]) > 0, "标题不能为空"
    # assert len(context["introduction"]) > 0 # 暂定不需要简介
    assert len(context["location"]) > 0, "地点不能为空"

    # 读取活动时间，检查合法性
    try:
        act_start = datetime.strptime(
            request.POST["lesson_start"], "%Y-%m-%d %H:%M")  # 活动开始时间
        act_end = datetime.strptime(
            request.POST["lesson_end"], "%Y-%m-%d %H:%M")  # 活动结束时间
        act_publish_day = {
            "instant": Course.PublishDay.instant,
            "3": Course.PublishDay.threeday,
            "2": Course.PublishDay.twoday,
            "1": Course.PublishDay.oneday,
        }[request.POST.get("publish_day")]  # 活动发布提前日期

        if act_publish_day == Course.PublishDay.instant:
            act_publish_time = datetime.now() + timedelta(seconds=10)   # 活动发布时间，立即发布
        else:
            act_publish_time = datetime.strptime(
                request.POST["publish_time"], "%Y-%m-%d %H:%M")  # 活动发布时间，指定的发布时间
    except:
        raise AssertionError("活动时间非法")
    context["start"] = act_start
    context["end"] = act_end
    context["publish_day"] = act_publish_day
    context["publish_time"] = act_publish_time

    if request.POST["need_apply"] == "True":
        assert datetime.now() < context["start"] - \
            timedelta(hours=1), "需要报名的活动必须提前至少一小时发起"

    assert check_ac_time_course(act_start, act_end), "活动时间非法"

    # 默认需要签到
    context["need_checkin"] = True
    # 默认不需要报名
    context["need_apply"] = request.POST["need_apply"] == "True"
    context["post_type"] = str(request.POST.get("post_type", ""))
    return context


def create_single_course_activity(request: HttpRequest) -> Tuple[int, bool]:
    """
    创建单次课程活动，是create_activity的简化版
    错误提示通过AssertionError抛出

    :param request: 发起单次课程活动的请求
    :type request: HttpRequest
    :return: 课程活动id，True(成功创建)/False(相似活动已存在)
    :rtype: Tuple[int, bool]
    """
    context = course_activity_base_check(request)

    # 获取组织和课程
    org = get_person_or_org(request.user, UTYPE_ORG)
    course = Course.objects.activated().get(organization=org)

    # 查找是否有类似活动存在
    old_ones = Activity.objects.activated().filter(
        title=context["title"],
        start=context["start"],
        # introduction=context["introduction"], # 暂定不需要简介
        location=context["location"],
        category=Activity.ActivityCategory.COURSE,  # 查重时要求是课程活动
    )
    if len(old_ones):
        assert len(old_ones) == 1, "创建活动时，已存在的相似活动不唯一"
        return old_ones[0].id, False

    # 获取默认审核老师
    examine_teacher = NaturalPerson.objects.get_teacher(
        APP_CONFIG.audit_teachers[0])

    # 获取活动所属课程的图片，用于viewActivity, examineActivity等页面展示
    image = str(course.photo)
    assert image, "获取课程图片失败"

    # 创建活动
    activity = Activity.objects.create(
        title=context["title"],
        organization_id=org,
        examine_teacher=examine_teacher,
        # introduction=context["introduction"],  # 暂定不需要简介
        location=context["location"],
        start=context["start"],
        end=context["end"],
        category=Activity.ActivityCategory.COURSE,
        need_checkin=True,  # 默认需要签到

        recorded=True,
        status=Activity.Status.UNPUBLISHED,
        publish_day=context["publish_day"],  # 发布提前天数
        publish_time=context["publish_time"],  # 发布时间
        need_apply=context["need_apply"]  # 是否需要报名

        # capacity, URL, bidding,
        # inner, end_before均为default
    )

    if context["need_apply"]:
        activity.endbefore = Activity.EndBefore.onehour
        activity.apply_end = activity.start - timedelta(hours=1)

    if not activity.need_apply:
        # 选课人员自动报名活动
        # 选课结束以后，活动参与人员从小组成员获取
        person_pos = list(Position.objects.activated().filter(
            org=course.organization).values_list("person", flat=True))
        if course.status == Course.Status.STAGE2:
            # 如果处于补退选阶段，活动参与人员从课程选课情况获取
            selected_person = list(CourseParticipant.objects.filter(
                course=course,
                status=CourseParticipant.Status.SUCCESS,
            ).values_list("person", flat=True))
            person_pos += selected_person
            person_pos = list(set(person_pos))
        members = NaturalPerson.objects.filter(
            id__in=person_pos)
        status = Participation.AttendStatus.APPLYSUCCESS
        create_participate_infos(activity, members, status=status)

        activity.current_participants = len(person_pos)
        activity.capacity = len(person_pos)
        activity.save()

    # 在活动发布时通知参与成员,创建定时任务并修改活动状态
    if activity.need_apply:
        ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.APPLYING}",
                      run_time=activity.publish_time)(activity.id, Activity.Status.UNPUBLISHED, Activity.Status.APPLYING)  # OK
        ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.WAITING}",
                      run_time=activity.start - timedelta(hours=1))(activity.id, Activity.Status.APPLYING, Activity.Status.WAITING)  # OK
    else:
        ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.WAITING}",
                      run_time=activity.publish_time)(activity.id, Activity.Status.UNPUBLISHED, Activity.Status.WAITING)  # OK

    ScheduleAdder(notifyActivity, id=f"activity_{activity.id}_newCourseActivity",
                  run_time=activity.publish_time)(activity.id, "newCourseActivity")  # OK

    # 引入定时任务：提前15min提醒、活动状态由WAITING变PROGRESSING再变END
    ScheduleAdder(notifyActivity, id=f"activity_{activity.id}_remind",
                  run_time=activity.start - timedelta(minutes=15))(activity.id, "remind")  # OK
    ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.PROGRESSING}",
                  run_time=activity.start)(activity.id, Activity.Status.WAITING, Activity.Status.PROGRESSING)
    ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.END}",
                  run_time=activity.end)(activity.id, Activity.Status.PROGRESSING, Activity.Status.END)
    activity.save()

    # 设置活动照片
    ActivityPhoto.objects.create(
        image=image, type=ActivityPhoto.PhotoType.ANNOUNCE, activity=activity)

    # 通知审核老师
    notification_create(
        receiver=examine_teacher.person_id,
        sender=request.user,
        typename=Notification.Type.NEEDDO,
        title=Notification.Title.VERIFY_INFORM,
        content="您有一个单次课程活动待审批",
        URL=f"/examineActivity/{activity.id}",
        relate_instance=activity,
        to_wechat=dict(app=WechatApp.AUDIT),
    )

    return activity.id, True


def modify_course_activity(request: HttpRequest, activity: Activity):
    """
    修改单次课程活动信息，是modify_activity的简化版
    错误提示通过AssertionError抛出

    :param request: 修改单次课程活动的请求
    :type request: HttpRequest
    :param activity: 待修改的活动
    :type activity: Activity
    """
    # 课程活动仅在待发布状态下可以修改
    assert activity.status == Activity.Status.UNPUBLISHED, \
        "课程活动只有在待发布状态才能修改。"

    context = course_activity_base_check(request)

    # 记录旧信息（以便发通知），写入新信息
    old_title = activity.title
    activity.title = context["title"]
    # activity.introduction = context["introduction"]# 暂定不需要简介
    old_location = activity.location
    activity.location = context["location"]
    old_start = activity.start
    activity.start = context["start"]
    old_end = activity.end
    activity.end = context["end"]
    old_publish_day = activity.publish_day
    activity.publish_day = context["publish_day"]
    old_publish_time = activity.publish_time
    activity.publish_time = context["publish_time"]
    old_need_apply = activity.need_apply
    activity.need_apply = context["need_apply"]

    if context["need_apply"]:
        activity.endbefore = Activity.EndBefore.onehour
        activity.apply_end = activity.start - timedelta(hours=1)

    activity.save()

    # 修改所有该时段的时间、地点
    if context["post_type"] == "modify_all" and activity.course_time is not None:
        course_time = CourseTime.objects.select_for_update().get(
            id=activity.course_time.id)
        course = course_time.course
        schedule_start = course_time.start
        schedule_end = course_time.end
        # 设置CourseTime初始时间为对应的 周几:hour:minute:second
        # 设置周几
        schedule_start += timedelta(
            days=(context["start"].weekday() - schedule_start.weekday()))
        schedule_end += timedelta(
            days=(context["end"].weekday() - schedule_end.weekday()))
        # 设置每周上课时间：hour:minute:second
        schedule_start = schedule_start.replace(hour=context["start"].hour,
                                                minute=context["start"].minute,
                                                second=context["start"].second)
        schedule_end = schedule_end.replace(hour=context["end"].hour,
                                            minute=context["end"].minute,
                                            second=context["end"].second)
        course_time.start = schedule_start
        course_time.end = schedule_end
        # 设置地点
        course.classroom = context["location"]
        course.need_apply = context["need_apply"]
        course.publish_day = context["publish_day"]
        course.save()
        course_time.save()
    # 目前只要编辑了活动信息，无论活动处于什么状态，都通知全体选课同学
    # if activity.status != Activity.Status.APPLYING and activity.status != Activity.Status.WAITING:
    #     return

    # 发布前参与同学无法获取课程信息，因此不需要发送通知
    # to_participants = [f"您参与的书院课程活动{old_title}发生变化"]
    # if old_title != activity.title:
    #     to_participants.append(f"活动更名为{activity.title}")
    # if old_location != activity.location:
    #     to_participants.append(f"活动地点修改为{activity.location}")
    # if old_start != activity.start:
    #     to_participants.append(
    #         f"活动开始时间调整为{activity.start.strftime('%Y-%m-%d %H:%M')}")

    # 更新定时任务
    if old_need_apply:
        # 删除报名中的状态阶段
        remove_job(job_id=f"activity_{activity.id}_{Activity.Status.APPLYING}")

    if activity.need_apply:
        ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.APPLYING}",
                      run_time=activity.publish_time)(activity.id, Activity.Status.UNPUBLISHED, Activity.Status.APPLYING)  # OK
        ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.WAITING}",
                      run_time=activity.start - timedelta(hours=1))(activity.id, Activity.Status.APPLYING, Activity.Status.WAITING)  # OK
    else:
        ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.WAITING}",
                      run_time=activity.publish_time)(activity.id, Activity.Status.UNPUBLISHED, Activity.Status.WAITING)  # OK

    ScheduleAdder(notifyActivity, id=f"activity_{activity.id}_newCourseActivity",
                  run_time=activity.publish_time)(activity.id, "newCourseActivity")  # OK
    ScheduleAdder(notifyActivity, id=f"activity_{activity.id}_remind",
                  run_time=activity.start - timedelta(minutes=15))(activity.id, "remind")  # OK
    ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.PROGRESSING}",
                  run_time=activity.start)(activity.id, Activity.Status.WAITING, Activity.Status.PROGRESSING)  # OK
    ScheduleAdder(changeActivityStatus, id=f"activity_{activity.id}_{Activity.Status.END}",
                  run_time=activity.end)(activity.id, Activity.Status.PROGRESSING, Activity.Status.END)  # OK

    # 发通知
    # notifyActivity(activity.id, "modification_par", "\n".join(to_participants))


def cancel_course_activity(request: HttpRequest, activity: Activity, cancel_all: bool = False):
    """
    取消课程活动，是cancel_activity的简化版，在聚合页面被调用

    在聚合页面中，应确保activity是课程活动，并且应检查activity.status，
    如果不是WAITING或PROGRESSING，不应调用本函数

    成功无返回值，失败返回错误消息
    （或者也可以在聚合页面判断出来能不能取消）

    :param request: 取消单次课程活动的请求
    :type request: HttpRequest
    :param activity: 待取消的活动
    :type activity: Activity
    :param cancel_all: 取消该时段所有活动, defaults to False
    :type cancel_all: bool, optional
    :return: 取消失败的话返回错误信息
    :rtype: string
    """
    # 只有UNPUBLISHED,WAITING和PROGRESSING允许取消
    if activity.status not in [
        Activity.Status.UNPUBLISHED,
        Activity.Status.WAITING,
        Activity.Status.PROGRESSING,
    ]:
        return f"课程活动状态为{activity.get_status_display()}，不可取消。"

    # 课程活动已于一天前开始则不能取消，这一点也可以在聚合页面进行判断
    if activity.status == Activity.Status.PROGRESSING:
        if activity.start.day != datetime.now().day:
            return "课程活动已于一天前开始，不能取消。"

    # 取消活动
    activity.status = Activity.Status.CANCELED
    # 目前只要取消了活动信息，无论活动处于什么状态，都通知全体选课同学
    notifyActivity(activity.id, "modification_par",
                   f"您报名的书院课程活动{activity.title}已取消（活动原定开始于{activity.start.strftime('%Y-%m-%d %H:%M')}）。")

    # 删除老师的审核通知（如果有）
    notification = Notification.objects.get(
        relate_instance=activity,
        typename=Notification.Type.NEEDDO
    )
    notification_status_change(notification, Notification.Status.DELETE)

    # 取消定时任务（需要先判断一下是否已经被执行了）
    if activity.start - timedelta(minutes=15) > datetime.now():
        remove_job(f"activity_{activity.id}_remind")
    if activity.start > datetime.now():
        remove_job(f"activity_{activity.id}_{Activity.Status.PROGRESSING}")
    if activity.end > datetime.now():
        remove_job(f"activity_{activity.id}_{Activity.Status.END}")

    activity.save()

    # 取消该时段所有活动！
    if cancel_all:
        # 设置结束 若cur_week >= end_week 则每周定时任务无需执行
        activity.course_time.end_week = activity.course_time.cur_week
        activity.course_time.save()


def remaining_willingness_point(user: NaturalPerson) -> int:
    """
    计算用户剩余的意愿点

    :param user: 当前用户
    :type user: NaturalPerson
    :raises NotImplementedError: 暂不启用
    :return: 剩余的意愿点值
    :rtype: int
    """
    raise NotImplementedError("暂时不使用意愿点")
    # 当前用户已经预选的课
    # 由于participant可能为空，重新启用的时候注意修改代码
    courses = Course.objects.filter(
        participant_set__person=user,
        participant_set__status=CourseParticipant.Status.SELECT)

    initial_point = 99  # 初始意愿点
    cost_point = courses.aggregate(Sum('bidding'))  # 已经使用的意愿点

    if cost_point:
        # cost_point可能为None
        return initial_point - cost_point['bidding__sum']
    else:
        return initial_point


def registration_status_check(course_status: Course.Status,
                              cur_status: CourseParticipant.Status,
                              to_status: CourseParticipant.Status) -> None:
    """
    检查选课状态的变化是否合法

    1. 预选阶段允许的状态变化: SELECT <-> UNSELECT
    2. 补退选阶段允许的状态变化: SUCCESS -> UNSELECT; FAILED -> SUCCESS; UNSELECT -> SUCCESS  

    异常: 抛出AssertionError，在调用处解决

    :param course_status: 课程所处的选课阶段
    :type course_status: Course.Status
    :param cur_status: 当前选课状态
    :type cur_status: CourseParticipant.Status
    :param to_status: 希望转变为的选课状态
    :type to_status: CourseParticipant.Status
    """

    if course_status == Course.Status.STAGE1:
        assert ((cur_status == CourseParticipant.Status.SELECT
                 and to_status == CourseParticipant.Status.UNSELECT)
                or (cur_status == CourseParticipant.Status.UNSELECT
                    and to_status == CourseParticipant.Status.SELECT))
    else:
        assert ((cur_status == CourseParticipant.Status.SUCCESS
                 and to_status == CourseParticipant.Status.UNSELECT)
                or (cur_status == CourseParticipant.Status.FAILED
                    and to_status == CourseParticipant.Status.SUCCESS)
                or (cur_status == CourseParticipant.Status.UNSELECT
                    and to_status == CourseParticipant.Status.SUCCESS))


def check_course_time_conflict(current_course: Course,
                               user: NaturalPerson) -> Tuple[bool, str]:
    """
    检查当前选择课程的时间和已选课程是否冲突

    :param current_course: 用户当前想选的课程
    :type current_course: Course
    :param user: 当前用户
    :type user: NaturalPerson
    :return: 是否冲突、发生冲突的具体原因
    :rtype: Tuple[bool, str]
    """
    selected_courses = Course.objects.selected(
        user, unfailed=True).prefetch_related("time_set")

    def time_hash(time: datetime):
        return time.weekday() * 1440 + time.hour * 60 + time.minute

    # 因为选择的课最多只能有6门，所以暂时用暴力算法
    for current_course_time in current_course.time_set.all():

        # 当前选择课程的上课时间
        current_start_time = current_course_time.start
        current_end_time = current_course_time.end

        for course in selected_courses:
            for course_time in course.time_set.all():
                start_time = course_time.start
                end_time = course_time.end

                # 效率不高，有待改进
                if not (time_hash(current_start_time) >= time_hash(end_time) or
                        time_hash(current_end_time) <= time_hash(start_time)):
                    # 发生冲突
                    return True, \
                        f"《{current_course.name}》和《{course.name}》的上课时间发生冲突！"

    # 没有冲突
    return False, ""
    '''
    # 循环较少的写法
    from django.db.models import Q
    conflict_course_names = set()
    for current_course_time in current_course.time_set.all():
        # 冲突时间
        conflict_times = CourseTime.objects.filter(
            # 已选的课程
            Q(course__in=selected_courses),
            # 开始比当前的结束时间早
            (Q(start__week_day=current_course_time.end.weekday() + 1,
               start__time__lte=current_course_time.end.time())
             | Q(start__week_day__lt=current_course_time.end.weekday() + 1))
            # 结束比当前的开始时间晚
            & (Q(end__week_day=current_course_time.start.weekday() + 1,
                 end__time__gte=current_course_time.start.time())
               | Q(end__week_day__gt=current_course_time.start.weekday() + 1))
        )
        if conflict_times.exists():
            # return True, f'《{conflict_times.first().course.name}》'
            conflict_course_names.union(
                conflict_times.values_list('course__name', flat=True))

    conflict_count = len(conflict_course_names)
    # 有冲突
    if conflict_count:
        return conflict_count, f'《{"》《".join(conflict_course_names)}》'
    # 没有冲突
    return conflict_count, ""
    '''


@logger.secure_func()
@transaction.atomic
def registration_status_change(course_id: int, user: NaturalPerson,
                               action: str) -> MESSAGECONTEXT:
    """
    学生点击选课或者取消选课后，更改学生的选课状态

    :param course_id: 当前课程的编号
    :type course_id: int
    :param user: 当前用户
    :type user: NaturalPerson
    :param action: 希望进行的操作，可能为"select"或"cancel"
    :type action: str
    :return: 操作是否成功执行
    :rtype: MESSAGECONTEXT
    """
    context = wrong("在修改选课状态的过程中发生错误，请联系管理员！")

    # 如果不把 user 锁起来，前面做的检查到后面更新数据库时可能已经无效了，会让用户选上超过6门或者时间冲突的课。
    # 最后get是为了强制对QuerySet求值，起到上锁的效果
    NaturalPerson.objects.select_for_update().get(id=user.id)

    # 在外部保证课程ID是存在的
    course = Course.objects.get(id=course_id)
    course_status = course.status

    if (course_status != Course.Status.STAGE1
            and course_status != Course.Status.STAGE2):
        return wrong("在非选课阶段不能选课！")

    if action == "select":
        if CourseParticipant.objects.filter(course_id=course_id,
                                            person=user).exists():
            participant_info = CourseParticipant.objects.get(
                course_id=course_id, person=user)
            cur_status = participant_info.status
        else:
            cur_status = CourseParticipant.Status.UNSELECT

        if course_status == Course.Status.STAGE1:
            to_status = CourseParticipant.Status.SELECT
        else:
            to_status = CourseParticipant.Status.SUCCESS

        # 选课不能超过6门
        if Course.objects.selected(user, unfailed=True).count() >= 6:
            return wrong("每位同学同时预选或选上的课程数最多为6门！")

        # 检查选课时间是否冲突
        is_conflict, message = check_course_time_conflict(course, user)

        if is_conflict:
            return wrong(message)
            # return wrong(f'与{is_conflict}门已选课程时间冲突: {message}')

        # 解锁成就-首次报名书院课程
        unlock_achievement(user, '首次报名书院课程')

    else:
        # action为取消预选或退选
        to_status = CourseParticipant.Status.UNSELECT

        # 不允许状态不存在，除非发生了严重的错误
        try:
            participant_info = CourseParticipant.objects.get(
                course_id=course_id, person=user)
            cur_status = participant_info.status
        except:
            return context

    # 检查当前选课状态、选课阶段和操作的一致性
    try:
        registration_status_check(course_status, cur_status, to_status)
    except AssertionError:
        return wrong("非法的选课状态修改！")

    # 暂时不使用意愿点选课
    # if (course_status == Course.Status.STAGE1
    #         and remaining_willingness_point(user) < course.bidding):
    #     context["warn_message"] = "剩余意愿点不足"

    # 更新选课状态
    try:
        with transaction.atomic():
            if to_status == CourseParticipant.Status.UNSELECT:
                Course.objects.filter(id=course_id).select_for_update().update(
                    current_participants=F("current_participants") - 1)
                CourseParticipant.objects.filter(course_id=course_id,
                                                 person=user).delete()
                succeed("成功取消选课！", context)
            else:
                # 处理并发问题
                course = Course.objects.select_for_update().get(id=course_id)
                if (course_status == Course.Status.STAGE2
                        and course.current_participants >= course.capacity):
                    wrong("选课人数已满！", context)
                else:
                    course.current_participants += 1
                    course.save()

                    CourseParticipant.objects.update_or_create(
                        course_id = course_id,
                        person = user,
                        defaults = {"status": to_status}
                    )
                    succeed("选课成功！", context)
    except:
        return context
    return context


def process_time(start: datetime, end: datetime) -> str:
    """
    把datetime对象转换成可读的时间表示

    :param start: 课程的开始时间
    :type start: datetime
    :param end: 课程的结束时间
    :type end: datetime
    :return: 可读的时间表示
    :rtype: str
    """
    chinese_display = ["一", "二", "三", "四", "五", "六", "日"]
    start_time = start.strftime("%H:%M")
    end_time = end.strftime("%H:%M")
    return f"周{chinese_display[start.weekday()]} {start_time}-{end_time}"


def course_to_display(courses: QuerySet[Course],
                      user: NaturalPerson,
                      detail: bool = False) -> List[dict]:
    """
    将课程信息转换为列表，方便前端呈现

    :param courses: 课程集合
    :type courses: QuerySet[Course]
    :param user: 当前用户
    :type user: NaturalPerson
    :param detail: 是否显示课程的详细信息, defaults to False
    :type detail: bool
    :return: 课程信息列表，用一个字典来传递课程的全部信息
    :rtype: List[dict]
    """
    display = []

    if detail:
        courses = courses.select_related('organization').prefetch_related(
            "time_set")
    else:
        # 预取，同时不查询不需要的字段
        courses = courses.defer(
            "classroom",
            "teacher",
            "introduction",
            "photo",
            "teaching_plan",
            "record_cal_method",
            "QRcode",
        ).select_related('organization').prefetch_related(
            Prefetch('participant_set',
                     queryset=CourseParticipant.objects.filter(person=user),
                     to_attr='participants'), "time_set")

    # 获取课程的基本信息
    for course in courses:
        course_info = {}

        # 选课页面和详情页面共用的信息
        course_info["name"] = course.name
        course_info["type"] = course.get_type_display()  # 课程类型
        course_info["avatar_path"] = course.organization.get_user_ava()

        course_time = []
        for time in course.time_set.all():
            course_time.append(process_time(time.start, time.end))
        course_info["time_set"] = course_time
        
        def linebreak(str):
            from re import sub
            return sub("((\r|\\\)+n)|((\r|\\\)+\n)", "\n", str)

        if detail:
            # 在课程详情页才展示的信息
            course_info["times"] = course.times  # 课程周数
            course_info["classroom"] = course.classroom
            course_info["teacher"] = course.teacher
            course_info["introduction"] = linebreak(course.introduction)
            course_info["teaching_plan"] = linebreak(course.teaching_plan)
            course_info["record_cal_method"] = linebreak(course.record_cal_method)
            course_info["organization_name"] = course.organization.oname
            course_info["have_QRcode"] = bool(course.QRcode)
            course_info["photo_path"] = course.get_photo_path()
            course_info["QRcode"] = course.get_QRcode_path()
            display.append(course_info)
            continue

        course_info["course_id"] = course.id
        course_info["capacity"] = course.capacity
        course_info["current_participants"] = course.current_participants
        course_info["status"] = course.get_status_display()  # 课程所处的选课阶段

        # 暂时不启用意愿点机制
        # course_info["bidding"] = int(course.bidding)

        # 当前学生的选课状态（注：course.participants是一个list）
        if course.participants:
            course_info["student_status"] = course.participants[
                0].get_status_display()
        else:
            course_info["student_status"] = "未选课"

        display.append(course_info)

    return display


@logger.secure_func()
def draw_lots():
    """
    等额抽签选出成功选课的学生，并修改学生的选课状态
    """
    courses = Course.objects.activated().filter(status=Course.Status.DRAWING)
    for course in courses:
        with transaction.atomic():
            participants = CourseParticipant.objects.filter(
                course=course, status=CourseParticipant.Status.SELECT)

            participants_num = participants.count()
            if participants_num <= 0:
                continue

            participants_id = list(participants.values_list("id", flat=True))
            capacity = course.capacity

            if participants_num <= capacity:
                # 选课人数少于课程容量，不用抽签
                CourseParticipant.objects.filter(
                    id__in=participants_id).select_for_update().update(
                        status=CourseParticipant.Status.SUCCESS)
                Course.objects.filter(id=course.id).select_for_update().update(
                    current_participants=participants_num)
            else:
                # 抽签；可能实现得有一些麻烦
                lucky_ones = sample(participants_id, capacity)
                unlucky_ones = list(
                    set(participants_id).difference(set(lucky_ones)))
                # 不确定是否要加悲观锁
                CourseParticipant.objects.filter(
                    id__in=lucky_ones).select_for_update().update(
                        status=CourseParticipant.Status.SUCCESS)
                CourseParticipant.objects.filter(
                    id__in=unlucky_ones).select_for_update().update(
                        status=CourseParticipant.Status.FAILED)
                Course.objects.filter(id=course.id).select_for_update().update(
                    current_participants=capacity)

            # 给选课成功的同学发送通知
            receivers = SQ.qsvlist(CourseParticipant.objects.filter(
                course=course,
                status=CourseParticipant.Status.SUCCESS,
            ), CourseParticipant.person, NaturalPerson.person_id)
            receivers = User.objects.filter(id__in=receivers)
            sender = course.organization.get_user()
            typename = Notification.Type.NEEDREAD
            title = Notification.Title.ACTIVITY_INFORM
            content = f"您好！您已成功选上课程《{course.name}》！"

            # 课程详情页面
            URL = f"/viewCourse/?courseid={course.id}"

            # 批量发送通知
            bulk_notification_create(
                receivers=receivers,
                sender=sender,
                typename=typename,
                title=title,
                content=content,
                URL=URL,
                to_wechat=dict(app=WechatApp.TO_PARTICIPANT,
                               level=WechatMessageLevel.IMPORTANT)
            )

            # 给选课失败的同学发送通知

            receivers = SQ.qsvlist(CourseParticipant.objects.filter(
                course=course,
                status=CourseParticipant.Status.FAILED,
            ), CourseParticipant.person, NaturalPerson.person_id)
            receivers = User.objects.filter(id__in=receivers)
            content = f"很抱歉通知您，您未选上课程《{course.name}》。"
            if len(receivers) > 0:
                bulk_notification_create(
                    receivers=receivers,
                    sender=sender,
                    typename=typename,
                    title=title,
                    content=content,
                    URL=URL,
                    to_wechat=dict(app=WechatApp.TO_PARTICIPANT,
                                   level=WechatMessageLevel.IMPORTANT),
                )


@logger.secure_func()
def change_course_status(cur_status: Course.Status, to_status: Course.Status) -> None:
    """
    作为定时任务，在课程设定的时间改变课程的选课阶段

    example: 
    scheduler.add_job(change_course_status, "date", 
                      id=f"course_{course_id}_{to_status}, run_date, args)

    :param cur_status: 课程的当前选课阶段
    :type cur_status: Course.Status
    :param to_status: 希望课程转变到的选课阶段
    :type to_status: Course.Status
    :raises AssertionError: 选课已经结束 / 两个状态间不匹配
    :raises AssertionError: 未提供当前阶段的信息
    """
    # 以下进行状态的合法性检查
    if cur_status is not None:
        if cur_status == Course.Status.WAITING:
            assert to_status == Course.Status.STAGE1, \
                f"不能从{cur_status}变更到{to_status}"
        elif cur_status == Course.Status.STAGE1:
            assert to_status == Course.Status.DRAWING, \
                f"不能从{cur_status}变更到{to_status}"
        elif cur_status == Course.Status.DRAWING:
            assert to_status == Course.Status.STAGE2, \
                f"不能从{cur_status}变更到{to_status}"
        elif cur_status == Course.Status.STAGE2:
            assert to_status == Course.Status.SELECT_END, \
                f"不能从{cur_status}变更到{to_status}"
        else:
            raise AssertionError("选课已经结束，不能再变化状态")
    else:
        raise AssertionError("未提供当前状态，不允许进行选课状态修改")
    courses = Course.objects.activated().filter(status=cur_status)
    if to_status == Course.Status.SELECT_END:
        courses = courses.select_related('organization')
    with transaction.atomic():
        for course in courses:
            if to_status == Course.Status.SELECT_END:
                # 选课结束，将选课成功的同学批量加入小组
                participants = CourseParticipant.objects.filter(
                    course=course,
                    status=CourseParticipant.Status.SUCCESS).select_related(
                        'person')
                organization = course.organization
                positions = []
                for participant in participants:
                    # 如果已有当前的离职状态，改成在职成员
                    Position.objects.current().filter(
                        person=participant.person,
                        org=organization,
                        status=Position.Status.DEPART,
                    ).update(pos=10,
                             is_admin=False,
                             semester=GLOBAL_CONFIG.semester,
                             status=Position.Status.INSERVICE)
                    # 检查是否已经加入小组
                    if not Position.objects.activated().filter(
                            person=participant.person,
                            org=organization).exists():
                        position = Position(person=participant.person,
                                            org=organization,
                                            semester=GLOBAL_CONFIG.semester)

                        positions.append(position)
                if positions:
                    with transaction.atomic():
                        Position.objects.bulk_create(positions)
        # 更新目标状态
        courses.select_for_update().update(status=to_status)


@logger.secure_func()
def register_selection(wait_for: timedelta | None = None):
    """
    添加定时任务，实现课程状态转变，每次发起课程时调用
    """
    # 预选和补退选的开始和结束时间
    year = GLOBAL_CONFIG.acadamic_year
    semester = GLOBAL_CONFIG.semester.value
    now = datetime.now()
    if wait_for is not None:
        now += wait_for
    stage1_start = str_to_time(APP_CONFIG.yx_election_start)
    stage1_start = max(stage1_start, now + timedelta(seconds=5))
    stage1_end = str_to_time(APP_CONFIG.yx_election_end)
    stage1_end = max(stage1_end, now + timedelta(seconds=10))
    publish_time = str_to_time(APP_CONFIG.publish_time)
    publish_time = max(publish_time, now + timedelta(seconds=15))
    stage2_start = str_to_time(APP_CONFIG.btx_election_start)
    stage2_start = max(stage2_start, now + timedelta(seconds=20))
    stage2_end = str_to_time(APP_CONFIG.btx_election_end)
    stage2_end = max(stage2_end, now + timedelta(seconds=25))
    # 定时任务：修改课程状态
    adder = MultipleAdder(change_course_status)
    adder.schedule(f'{year}_{semester}_选课_stage1_start',
                   run_time=stage1_start)(Course.Status.WAITING, Course.Status.STAGE1)
    adder.schedule(f'{year}_{semester}_选课_stage1_end',
                   run_time=stage1_end)(Course.Status.STAGE1, Course.Status.DRAWING)
    ScheduleAdder(draw_lots, id=f'{year}_{semester}_选课_publish',
                  run_time=publish_time)()
    adder.schedule(f'{year}_{semester}_选课_stage2_start',
                   run_time=stage2_start)(Course.Status.DRAWING, Course.Status.STAGE2)
    adder.schedule(f'{year}_{semester}_选课_stage2_end',
                   run_time=stage2_end)(Course.Status.STAGE2, Course.Status.SELECT_END)
    # 状态随时间的变化: WAITING-STAGE1-WAITING-STAGE2-END


def course_base_check(request, if_new=None):
    """
    选课单变量合法性检查并准备变量
    """
    context = dict()
    # 字符串字段合法性检查
    try:
        # name, introduction, classroom 创建时不能为空
        context = my_messages.read_content(
            request.POST,
            "name",
            'teacher',
            "introduction",
            "classroom",
            "teaching_plan",
            "hours_per_class",
            "record_cal_method",
            "need_apply",
            "publish_day",
            _trans_func=str,
            _default="",
        )
        assert len(context["name"]) > 0, "课程名称不能为空！"
        assert len(context["introduction"]) > 0, "课程介绍不能为空！"
        assert len(context["teaching_plan"]) > 0, "教学计划不能为空！"
        assert len(context["record_cal_method"]) > 0, "学时计算方法不能为空！"
        assert len(context["classroom"]) > 0, "上课地点不能为空！"
        assert context["need_apply"] in ["True", "False"], "是否需要报名必须为给定值！"
        assert context["publish_day"] in [
            "instant", "1", "2", "3"], "信息发布时间必须为给定值！"
    except Exception as e:
        return wrong(str(e))
    # int类型合法性检查

    type_num = request.POST.get("type", "")  # 课程类型
    capacity = request.POST.get("capacity", -1)
    # context['times'] = int(request.POST["times"])    #课程上课周数
    try:
        cur_info = "记得选择课程类型哦！"
        type_num = int(type_num)
        cur_info = "课程类型仅包括德智体美劳五种！"
        assert 0 <= type_num < 5
        cur_info = "课程容量应当大于0！"
        assert int(capacity) > 0
    except:
        return wrong(cur_info)
    context['type'] = type_num
    context['capacity'] = capacity

    # 图片类型合法性检查
    try:
        announcephoto = request.FILES.get("photo")
        pic = None
        if announcephoto:
            pic = announcephoto
            assert if_image(pic) == 2, "课程预告图片文件类型错误！"
        else:
            for i in range(5):
                if request.POST.get(f'picture{i+1}'):
                    pic = request.POST.get(f'picture{i+1}')
        context["photo"] = pic
        context["QRcode"] = request.FILES.get("QRcode")
        if if_new is None:
            assert context["photo"] is not None, "缺少课程预告图片！"
        assert if_image(context["QRcode"]) != 1, "微信群二维码图片文件类型错误！"
    except Exception as e:
        return wrong(str(e))

    # 每周课程时间合法性检查
    # TODO: 需要增加是否可以修改时间的安全性检查
    course_starts = request.POST.getlist("start")
    course_ends = request.POST.getlist("end")
    course_starts = [
        datetime.strptime(course_start, "%Y-%m-%d %H:%M")
        for course_start in course_starts
        if course_start != ''
    ]
    course_ends = [
        datetime.strptime(course_end, "%Y-%m-%d %H:%M")
        for course_end in course_ends
        if course_end != ''
    ]
    try:
        for i in range(len(course_starts)):
            assert check_ac_time_course(
                course_starts[i], course_ends[i]), f'第{i+1}次上课时间起止时间有误！'
            # 课程每周同一次课的开始和结束时间应当处于同一天
            assert course_starts[i].date(
            ) == course_ends[i].date(), f'第{i+1}次上课起止时间应当为同一天'
    except Exception as e:
        return wrong(str(e))
    context['course_starts'] = course_starts
    context['course_ends'] = course_ends
    context['publish_day'] = {
        "instant": Course.PublishDay.instant,
        "3": Course.PublishDay.threeday,
        "2": Course.PublishDay.twoday,
        "1": Course.PublishDay.oneday,
    }[context['publish_day']]
    org = get_person_or_org(request.user, UTYPE_ORG)
    context['organization'] = org

    succeed("合法性检查通过！", context)
    return context


def create_course(request, course_id=None):
    '''
    检查课程，合法时寻找该课程，不存在时创建
    返回(course.id, created)
    '''
    context = dict()

    try:
        context = course_base_check(request, course_id)
        if context["warn_code"] == 1:  # 合法性检查出错！
            return context
    except:
        return wrong("检查参数合法性时遇到不可预料的错误。如有需要，请联系管理员解决!")

    # 编辑已有课程
    if course_id is not None:
        try:
            course = Course.objects.get(id=int(course_id))
            with transaction.atomic():
                if course.status in [Course.Status.WAITING]:
                    course_time = course.time_set.all()
                    course_time.delete()
                course.name = context["name"]
                course.classroom = context["classroom"]
                course.teacher = context['teacher']
                course.introduction = context["introduction"]
                course.teaching_plan = context["teaching_plan"]
                course.hours_per_class = context["hours_per_class"]
                course.record_cal_method = context["record_cal_method"]
                course.type = context['type']
                course.capacity = context["capacity"]
                course.need_apply = context["need_apply"]
                course.publish_day = context["publish_day"]
                course.photo = context['photo'] if context['photo'] is not None else course.photo
                if context['QRcode']:
                    course.QRcode = context["QRcode"]
                course.save()

                if course.status in [Course.Status.WAITING]:
                    for i in range(len(context['course_starts'])):
                        CourseTime.objects.create(
                            course=course,
                            start=context['course_starts'][i],
                            end=context['course_ends'][i],
                        )
        except:
            return wrong("修改课程时遇到不可预料的错误。如有需要，请联系管理员解决!")
        context["cid"] = course_id
        context["warn_code"] = 2
        context["warn_message"] = "修改课程成功！"
    # 创建新课程
    else:
        try:
            with transaction.atomic():
                course = Course.objects.create(
                    name=context["name"],
                    organization=context['organization'],
                    classroom=context["classroom"],
                    teacher=context['teacher'],
                    introduction=context["introduction"],
                    teaching_plan=context["teaching_plan"],
                    record_cal_method=context["record_cal_method"],
                    type=context['type'],
                    capacity=context["capacity"],
                    need_apply=context["need_apply"],
                    publish_day=context["publish_day"]
                )
                course.photo = context['photo']
                if context['QRcode']:
                    course.QRcode = context["QRcode"]
                course.save()

                for i in range(len(context['course_starts'])):
                    CourseTime.objects.create(
                        course=course,
                        start=context['course_starts'][i],
                        end=context['course_ends'][i],
                    )
            register_selection()  # 每次发起课程，创建定时任务
        except:
            return wrong("创建课程时遇到不可预料的错误。如有需要，请联系管理员解决!")
        context["cid"] = course.id
        context["warn_code"] = 2
        context["warn_message"] = "创建课程成功！"

    return context


def cal_participate_num(course: Course) -> dict:
    """
    计算该课程对应组织所有成员的参与次数
    return {Naturalperson.id:参与次数}
    前端使用的时候直接读取字典的值就好了
    :param course: 选择要计算的课程
    :type course: Course
    :return: 返回统计数据
    :rtype: dict
    """
    org = course.organization
    activities = Activity.objects.activated().filter(
        organization_id=org,
        status=Activity.Status.END,
        category=Activity.ActivityCategory.COURSE,
    )
    # 只有小组成员才可以有学时
    members = Position.objects.activated().filter(
        pos__gte=1,
        person__identity=NaturalPerson.Identity.STUDENT,
        org=org,
    ).values_list("person", flat=True)
    all_participants = SQ.qsvlist(
        Participation.objects.activated(no_unattend=True)
        .filter(SQ.mq(Participation.activity, IN=activities),
                SQ.mq(Participation.person, IN=members)),
        Participation.person)
    participate_num = dict(Counter(all_participants))
    # 没有参加的参与次数设置为0
    participate_num.update(
        {id: 0 for id in members if id not in participate_num})
    return participate_num


def check_post_and_modify(records: QuerySet[CourseRecord], post_data: QueryDict) -> MESSAGECONTEXT:
    """
    records和post_data分别为原先和更新后的list
    检查post表单是否可以为这个course对应的内容，
    如果可以，修改学时
    - 返回wrong|succeed
    - 不抛出异常
    在检查过程中，会使用事务和 select_for_update 来保证原子性。
    :param records: 原本的学时数据
    :type records: QuerySet[CourseRecord]
    :param post_data: 由前端上传上来的修改结果
    :type post_data: QueryDict
    :return: 检查结果
    :rtype: MESSAGECONTEXT
    """
    try:
        # 对每一条记录而言
        with transaction.atomic():
            records = records.select_for_update()
            for record in records:
                # 选取id作为匹配键
                key = str(record.person.id)
                assert key in post_data.keys(), "提交的人员信息不匹配，请联系管理员！"

                # 读取小时数
                bonus_hours = float(post_data.get(str(key), -1))
                record.bonus_hours = bonus_hours
                record.total_hours = bonus_hours + record.attend_times * record.hours_per_class
                assert record.total_hours >= 0, "总学时数据为负数，请检查输入数据！"
                # 更新是否有效
                record.invalid = (record.total_hours < APP_CONFIG.least_record_hours)

            CourseRecord.objects.bulk_update(records, ["bonus_hours", "total_hours", "invalid"])

        return succeed("修改学时信息成功！")
    except AssertionError as e:
        # 此时相当于出现用户应该知晓的信息
        return wrong(str(e))
    except IntegrityError:
        return wrong("数据库操作失败，请联系管理员！")
    except:
        return wrong("数据格式异常，请检查输入数据！")


def finish_course(course: Course):
    """
    结束课程
    设置课程状态为END 生成学时表并通知同学该课程已结束。
    """
    # 若存在课程活动未结束则无法结束课程。
    cur_activities = Activity.objects.activated().filter(
        organization_id=course.organization,
        category=Activity.ActivityCategory.COURSE).exclude(
            status__in=[
                Activity.Status.CANCELED,
                Activity.Status.END,
            ])
    if cur_activities.exists():
        return wrong("存在尚未结束的课程活动，请在所有课程活动结束以后再结束课程。")

    try:
        # 取消发布每周定时活动
        course_times = course.time_set
        for course_time in course_times.all():
            course_time.end_week = course_time.cur_week
            course_time.save()
    except:
        return wrong("取消课程每周定时活动时失败，请联系管理员！")
    try:
        # 生成学时表
        participate_num = cal_participate_num(course)
        participants = NaturalPerson.objects.activated().filter(
            id__in=participate_num.keys())
        course_record_list = []
        for participant in participants:
            # 如果存在相同学期的学时表，则不创建
            if not CourseRecord.objects.current().filter(
                    person=participant, course=course
            ).exists():
                hours_per_class = course.hours_per_class
                attend_times = participate_num[participant.id]
                total_hours = attend_times * hours_per_class
                course_record_list.append(CourseRecord(
                    person = participant,
                    course = course,
                    attend_times = attend_times,
                    hours_per_class = hours_per_class,
                    bonus_hours = 0.0,
                    total_hours = total_hours,
                    invalid = (total_hours < 8),
                ))
        CourseRecord.objects.bulk_create(course_record_list)
    except:
        return wrong("生成学时表失败，请联系管理员！")
    try:
        # 通知课程小组成员该课程已结束
        title = f'课程结束通知！'
        msg = f'{course.name}在本学期的课程已结束！'
        receivers = SQ.qsvlist(participants, NaturalPerson.person_id)
        receivers = User.objects.filter(id__in=receivers)
        bulk_notification_create(
            receivers=list(receivers),
            sender=course.organization.get_user(),
            typename=Notification.Type.NEEDREAD,
            title=title,
            content=msg,
            URL=f"/viewCourse/?courseid={course.id}",
            to_wechat=dict(app=WechatApp.TO_PARTICIPANT),
        )
        # 设置课程状态
    except:
        return wrong("生成通知失败，请联系管理员！")
    course.status = Course.Status.END
    course.save()
    return succeed("结束课程成功！")


def _excel_response(workbook: openpyxl.Workbook, file_name: str) -> HttpResponse:
    '''创建Excel文件回应，file_name为未转义的文件名(不含后缀)'''
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename={quote(file_name)}.xlsx'
    workbook.save(response)
    return response


def _write_detail_sheet(detail_sheet: openpyxl.worksheet.worksheet.Worksheet,
                        records: QuerySet[CourseRecord], title: str = '详情') -> None:
    '''将学时信息写入Excel文件的detail_sheet中'''
    # 注意，标题中的中文符号如：无法被解读
    detail_sheet.title = title
    # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
    detail_header = ['课程', '姓名', '学号', '次数', '额外学时', '总学时', '学年', '学期', '有效']
    detail_sheet.append(detail_header)
    _M = CourseRecord
    for record in records.values_list(
        SQ.f(_M.course, Course.name), SQ.f(_M.extra_name),
        SQ.f(_M.person, NaturalPerson.name),
        SQ.f(_M.person, NaturalPerson.person_id, User.username),
        SQ.f(_M.attend_times), SQ.f(_M.bonus_hours), SQ.f(_M.total_hours),
        SQ.f(_M.year), SQ.f(_M.semester), SQ.f(_M.invalid),
    ):
        record_info = [
            record[0] or record[1],
            *record[2:7],
            f'{record[7]}-{record[7] + 1}',
            '春' if record[8] == Semester.SPRING else '秋',
            '否' if record[9] else '是',
        ]
        # 将每一个对象的所有字段的信息写入一行内
        detail_sheet.append(record_info)


def download_course_record(course: Course = None, year: int = None, semester: Semester = None) -> HttpResponse:
    """返回需要导出的学时信息文件
    course:
        提供course时为单个课程服务，只导出该课程的相关人员的学时信息
        不提供时下载所有学时信息，注意，只有相关负责老师可以访问！
    :param course: 所选择的课程, defaults to None
    :type course: Course, optional
    :param year: 所选择的学年, defaults to None
    :type year: int, optional
    :param semester: 所选择的学期, defaults to None
    :type semester: Semester, optional
    :return: 返回下载的文件数据
    :rtype: HttpResponse
    """
    wb = openpyxl.Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    # 学时筛选内容
    filter_kws = {}
    if course is not None:
        filter_kws[SQ.f(CourseRecord.course)] = course
    if year is not None:
        filter_kws[SQ.f(CourseRecord.year)] = year
    if semester is not None:
        filter_kws[SQ.f(CourseRecord.semester)] = semester

    if course is not None:
        # 助教下载自己课程的学时
        records = CourseRecord.objects.filter(**filter_kws)
        _write_detail_sheet(wb.active, records)
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        return _excel_response(wb, f'{course.name}-{now}')

    # 老师下载所有课程的学时
    # 获取第一个工作表（detail_sheet）
    detail_sheet = wb.active
    # 设置明细和汇总两个sheet的相关信息
    total_sheet: openpyxl.worksheet.worksheet.Worksheet = wb.create_sheet('汇总', 0)
    first_line = ['学号', '姓名', '总有效学时', '总无效学时']
    first_line.extend(Course.CourseType.labels)
    first_line.append('其他')
    total_sheet.append(first_line)

    # 下载所有学时信息，包括无效学时
    all_person = NaturalPerson.objects.activated().filter(
        identity=NaturalPerson.Identity.STUDENT)

    # 汇总表信息，姓名，学号，总学时
    relate = SQ.Reverse(CourseRecord.person)
    person_record = all_person.annotate(
        record_hours=Sum(SQ.f(relate, CourseRecord.total_hours),
                            filter=SQ.mq(relate, invalid=False, **filter_kws)),
        invalid_hours=Sum(SQ.f(relate, CourseRecord.total_hours),
                            filter=SQ.mq(relate, invalid=True, **filter_kws)),
    ).order_by(SQ.f(NaturalPerson.person_id, User.username))

    def _sum_hours(records: QuerySet[CourseRecord]) -> float:
        agg = records.filter(**filter_kws).aggregate(sum=Sum('total_hours'))
        return agg['sum'] or 0

    for person in person_record.select_related(SQ.f(NaturalPerson.person_id)):
        line = [person.person_id.username, person.name,
                person.record_hours or 0, 
                person.invalid_hours or 0]
        valid_records = SQ.sfilter(CourseRecord.person, person).exclude(invalid=True)
        # 计算每个类别的学时
        for course_type in list(Course.CourseType):
            line.append(_sum_hours(valid_records.filter(course__type=course_type)))
        # 计算没有对应Course的学时
        line.append(_sum_hours(valid_records.filter(course__isnull=True)))
        total_sheet.append(line)

    # 详细信息
    records = SQ.mfilter(CourseRecord.person, IN=all_person).filter(**filter_kws)
    order = SQ.f(CourseRecord.person, NaturalPerson.person_id, User.username)
    _write_detail_sheet(detail_sheet, records.order_by(order))
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    return _excel_response(wb, f'学时汇总-{now}')


def download_select_info(single_course: Course | None = None):
    """
    下载选课信息
    single_course:
        提供single_course时为单个课程服务，只导出该课程的相关人员的选课信息
        不提供时下载所有选课信息，注意，此时只有相关负责老师可以访问！
        不提供手动选课人员名单。
    :return: 返回下载的文件数据
    :rtype: HttpResponse
    """
    wb = openpyxl.Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet_header = ['姓名', '学号']
    if single_course is not None:
        courses = [single_course]
    else:
        courses = Course.objects.activated()
    # 为每一门课创建一个新的sheet
    for course in courses:
        # 生成新的sheet，并设置表头
        sheet = wb.create_sheet(title=f'{course.name}')
        sheet.append(sheet_header)
        # 导出课程小组信息
        class_members = NaturalPerson.objects.filter(
            id__in=Position.objects.activated()
                      .filter(org=course.organization)
                      .values_list("person", flat=True)
        )
        for info in class_members.values_list(
            SQ.f(NaturalPerson.name), 
            SQ.f(NaturalPerson.person_id, User.username)
        ):
            person_info = [
                info[0],
                info[1]
            ]
            sheet.append(person_info)
    # 删除多余的第一个sheet
    wb.remove_sheet(wb.active)
    # 设置文件名
    semester = "春" if courses[0].semester == Semester.SPRING else "秋"
    year = (courses[0].year + 1) if semester == "春" else courses[0].year
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M')
    if single_course is not None:
        file_name = f'{year}{semester}{course.name}选课名单-{ctime}'
    else:
        file_name = f'{year}{semester}选课名单汇总-{ctime}'
    # 保存并返回
    return _excel_response(wb, file_name)
