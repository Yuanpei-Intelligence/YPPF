"""
course_views.py

选课页面: selectCourse
课程详情页面: viewCourse
"""
from app.views_dependency import *
from app.models import (
    NaturalPerson,
    Semester,
    Activity,
    Course,
    CourseRecord,
)
from app.course_utils import (
    cancel_course_activity,
    create_single_course_activity,
    modify_course_activity,
    registration_status_change,
    course_to_display,
    create_course,
    cal_participate_num,
    check_post_and_modify,
    finish_course,
    str_to_time,
)
from app.utils import get_person_or_org

from datetime import datetime

from django.db import transaction
from io import BytesIO
from openpyxl import Workbook
import re
from zhon.hanzi import punctuation

__all__ = [
    'editCourseActivity',
    'addSingleCourseActivity',
    'showCourseActivity',
    'showCourseRecord',
    'selectCourse',
    'viewCourse',
]


@login_required(redirect_field_name="origin")
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(EXCEPT_REDIRECT, source='course_views[editCourseActivity]', record_user=True)
def editCourseActivity(request, aid):
    """
    编辑单次书院课程活动，addActivity的简化版
    """
    # 检查用户身份
    try:
        valid, user_type, html_display = utils.check_user_type(request.user)
        # assert valid  已经在check_user_access检查过了
        me = utils.get_person_or_org(request.user, user_type)  # 这里的me应该为小组账户
        aid = int(aid)
        activity = Activity.objects.get(id=aid)
        if user_type == "Person":
            html_display = utils.user_login_org(
                request, activity.organization_id)
            if html_display['warn_code'] == 1:
                return redirect(message_url(wrong(html_display["warn_message"])))
            else:  # 成功以小组账号登陆
                # 防止后边有使用，因此需要赋值
                user_type = "Organization"
                request.user = activity.organization_id.organization_id  # 小组对应user
                me = activity.organization_id  # 小组
        if activity.organization_id != me:
            return redirect(message_url(wrong("无法修改其他课程小组的活动!")))
        html_display["is_myself"] = True
    except Exception as e:
        log.record_traceback(request, e)
        return EXCEPT_REDIRECT

    # 这个页面只能修改书院课程活动(category=1)
    if activity.category != Activity.ActivityCategory.COURSE:
        return redirect(message_url(wrong('当前活动不是书院课程活动!'),
                                    f'/viewActivity/{activity.id}'))
    # 课程活动无需报名，在开始前都是等待中的状态
    if activity.status != Activity.Status.WAITING:
        return redirect(message_url(wrong('当前活动状态不允许修改!'),
                                    f'/viewActivity/{activity.id}'))

    if request.method == "POST" and request.POST:
        # 修改活动
        try:
            # 只能修改自己的活动
            with transaction.atomic():
                activity = Activity.objects.select_for_update().get(id=aid)
                org = get_person_or_org(request.user, "Organization")
                assert activity.organization_id == org
                modify_course_activity(request, activity)
            html_display["warn_msg"] = "修改成功。"
            html_display["warn_code"] = 2
        except Exception as e:
            log.record_traceback(request, e)
            return EXCEPT_REDIRECT

    # 前端使用量
    html_display["applicant_name"] = me.oname
    html_display["app_avatar_path"] = me.get_user_ava()
    bar_display = utils.get_sidebar_and_navbar(request.user, "修改课程活动")

    # 前端使用量，均可编辑
    title = utils.escape_for_templates(activity.title)
    location = utils.escape_for_templates(activity.location)
    start = activity.start.strftime("%Y-%m-%d %H:%M")
    end = activity.end.strftime("%Y-%m-%d %H:%M")
    # introduction = escape_for_templates(activity.introduction) # 暂定不需要简介
    edit = True  # 前端据此区分是编辑还是创建

    # 判断本活动是否为长期定时活动
    course_time_tag = (activity.course_time is not None)

    return render(request, "lesson_add.html", locals())


@login_required(redirect_field_name="origin")
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(EXCEPT_REDIRECT, source='course_views[addSingleCourseActivity]', record_user=True)
def addSingleCourseActivity(request):
    """
    创建单次书院课程活动，addActivity的简化版
    """
    # 检查用户身份
    try:
        valid, user_type, html_display = utils.check_user_type(request.user)
        # assert valid  已经在check_user_access检查过了
        me = utils.get_person_or_org(request.user, user_type)  # 这里的me应该为小组账户
        if user_type != "Organization" or me.otype.otype_name != COURSE_TYPENAME:
            return redirect(message_url(wrong('书院课程小组账号才能开设课程活动!')))
        if me.oname == YQP_ONAME:
            return redirect("/showActivity")  # TODO: 可以重定向到书院课程聚合页面
        html_display["is_myself"] = True
    except Exception as e:
        log.record_traceback(request, e)
        return EXCEPT_REDIRECT

    if request.method == "POST" and request.POST:
        # 创建活动
        try:
            with transaction.atomic():
                aid, created = create_single_course_activity(request)
                if not created:
                    return redirect(message_url(
                        succeed('存在信息相同的课程活动，已为您自动跳转!'),
                        f'/viewActivity/{aid}'))
                return redirect(f"/editCourseActivity/{aid}")
        except Exception as e:
            log.record_traceback(request, e)
            return EXCEPT_REDIRECT

    # 前端使用量
    html_display["applicant_name"] = me.oname
    html_display["app_avatar_path"] = me.get_user_ava()
    bar_display = utils.get_sidebar_and_navbar(request.user, "发起单次课程活动")
    edit = False  # 前端据此区分是编辑还是创建
    course_time_tag = False

    return render(request, "lesson_add.html", locals())


@login_required(redirect_field_name='origin')
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(source='course_views[showCourseActivity]', record_user=True)
def showCourseActivity(request):
    """
    筛选本学期已结束的课程活动、未开始的课程活动，在课程活动聚合页面进行显示。
    """

    # Sanity check and start a html_display.
    _, user_type, html_display = utils.check_user_type(request.user)
    me = get_person_or_org(request.user, user_type)  # 获取自身

    if user_type != "Organization" or me.otype.otype_name != COURSE_TYPENAME:
        return redirect(message_url(wrong('只有书院课程组织才能查看此页面!')))
    my_messages.transfer_message_context(request.GET, html_display)

    all_activity_list = (
        Activity.objects
        .activated()
        .filter(organization_id=me)
        .filter(category=Activity.ActivityCategory.COURSE)
        .order_by("-start")
    )

    future_activity_list = (
        all_activity_list.filter(
            status__in=[
                Activity.Status.REVIEWING,
                Activity.Status.APPLYING,
                Activity.Status.WAITING,
                Activity.Status.PROGRESSING,
            ]
        )
    )

    finished_activity_list = (
        all_activity_list
        .filter(
            status__in=[
                Activity.Status.END,
                Activity.Status.CANCELED,
            ]
        )
        .order_by("-end")
    )  # 本学期的已结束活动（包括已取消的）

    bar_display = utils.get_sidebar_and_navbar(
        request.user, navbar_name="我的活动")

    # 取消单次活动
    if request.method == "POST" and request.POST:
        cancel_all = False
        # 获取待取消的活动
        try:
            aid = int(request.POST.get("cancel-action"))
            post_type = str(request.POST.get("post_type"))
            if post_type == "cancel_all":
                cancel_all = True
            activity = Activity.objects.get(id=aid)
        except:
            return redirect(message_url(wrong('遇到不可预料的错误。如有需要，请联系管理员解决!'), request.path))

        if activity.organization_id != me:
            return redirect(message_url(wrong('您没有取消该课程活动的权限!'), request.path))

        if activity.status in [
            Activity.Status.REJECT,
            Activity.Status.ABORT,
            Activity.Status.END,
            Activity.Status.CANCELED,
        ]:
            return redirect(message_url(wrong('该课程活动已结束，不可取消!'), request.path))

        assert activity.status not in [
            Activity.Status.REVIEWING,
            Activity.Status.APPLYING,
        ], "课程活动状态非法"  # 课程活动不应出现这两个状态

        # 取消活动
        with transaction.atomic():
            activity = Activity.objects.select_for_update().get(id=aid)
            error = cancel_course_activity(request, activity, cancel_all)

        # 无返回值表示取消成功，有则失败
        if error is None:
            html_display["warn_code"] = 2
            html_display["warn_message"] = "成功取消活动。"
        else:
            return redirect(message_url(wrong(error)), request.path)

    return render(request, "org_show_course_activity.html", locals())


@login_required(redirect_field_name="origin")
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(EXCEPT_REDIRECT, source='course_views[showCourseRecord]', record_user=True)
def showCourseRecord(request):
    '''
    展示及修改学时数据
    在开启修改功能前，显示本学期已完成的所有课程活动的学生的参与次数
    开启修改功能后，自动创建学时表，并且允许修改学时
    '''
    # ----身份检查----
    _, user_type, _ = utils.check_user_type(request.user)
    me = utils.get_person_or_org(request.user)  # 获取自身
    if user_type == "Person":
        return redirect(message_url(wrong('学生账号不能访问此界面！')))
    if me.otype.otype_name != COURSE_TYPENAME:
        return redirect(message_url(wrong('非书院课程组织账号不能访问此界面！')))

    # 提取课程，后端保证每个组织只有一个Course字段

    # 获取课程开设筛选信息
    year = CURRENT_ACADEMIC_YEAR
    semester = Semester.now()

    course = Course.objects.activated().filter(
        organization=me,
        year=year,
        semester=semester,
    )
    if len(course) == 0: # 尚未开课的情况
        return redirect(message_url(wrong('没有检测到该组织本学期开设的课程。')))
    # TODO: 报错 这是代码不应该出现的bug
    assert len(course) == 1, "检测到该组织的课程超过一门，属于不可预料的错误，请及时处理！"
    course = course.first()

    # 是否可以编辑
    editable = course.status == Course.Status.END
    # 获取前端可能的提示
    messages = my_messages.transfer_message_context(request.GET)

    # -------- POST 表单处理 --------
    # 默认状态为正常
    if request.method == "POST" and request.POST:
        if not editable:
            # 由于未开放修改功能时前端无法通过表格和按钮修改和提交，
            # 所以如果出现POST请求，则为非法情况
            post_type = str(request.POST.get("post_type"))
            if post_type == "end":
                with transaction.atomic():
                    course = Course.objects.select_for_update().get(id=course.id)
                    messages = finish_course(course)
                return redirect(message_url(messages, request.path))
            else:
                return redirect(message_url(
                    wrong('学时修改尚未开放。如有疑问，请联系管理员！'), request.path))
        # 导出学时为表格
        if request.POST.get("download_course_record") is not None:
            response = downloadCourseRecord(me,year,semester)
            return response
        # 不是其他post类型时的默认行为
        with transaction.atomic():
            # 检查信息并进行修改
            record_search = CourseRecord.objects.filter(
                course=course,
                year=year,
                semester=semester,
            ).select_for_update()
            messages = check_post_and_modify(record_search, request.POST)
            # TODO: 发送微信消息?不一定需要

    # -------- GET 部分 --------
    # 如果进入这个页面时课程的状态(Course.Status)为未结束，那么只能查看不能修改，此时从函数读取
    # 每次进入都获取形如{id: times}的字典，这里id是naturalperson的主键id而不是userid
    participate_raw = cal_participate_num(course)
    if not editable:
        convert_dict = participate_raw    # 转换为字典方便查询, 这里已经是字典了
        # 选取人选
        participant_list = NaturalPerson.objects.activated().filter(
            id__in=convert_dict.keys()
        )

        # 转换为前端使用的list
        records_list = [
            {
                "pk": person.id,
                "name": person.name,
                "grade": person.stu_grade,
                "avatar": person.get_user_ava(),
                "times": convert_dict[person.id],   # 参与次数
            } for person in participant_list
        ]

    # 否则可以修改表单，从CourseRecord读取
    else:

        records_list = []
        with transaction.atomic():
            # 查找此课程本学期所有成员的学时表
            record_search = CourseRecord.objects.filter(
                course=course,
                year=year,
                semester=semester,
            ).select_for_update().select_related(
                "person"
            )   # Prefetch person to use its name, stu_grade and avatar. Help speed up.

            # 前端循环list
            for record in record_search:
                # 每次都需要更新一下参与次数，避免出现手动调整签到但是未能记录在学时表的情况
                record.attend_times = participate_raw[record.person.id]
                records_list.append({
                    "pk": record.person.id,
                    "name": record.person.name,
                    "grade": record.person.stu_grade,
                    "avatar": record.person.get_user_ava(),
                    "times": record.attend_times,
                    "hours": record.total_hours
                })
            CourseRecord.objects.bulk_update(record_search, ["attend_times"])

    # 前端呈现信息，用于展示
    course_info = {
        'course': course.name,
        'year': year,
        'semester': "春季" if semester == Semester.SPRING else "秋季",
    }
    bar_display = utils.get_sidebar_and_navbar(request.user, "课程学时")

    render_context = dict(
        course_info=course_info, records_list=records_list,
        editable=editable,
        bar_display=bar_display, messages=messages,
    )
    return render(request, "course_record.html", render_context)


@login_required(redirect_field_name="origin")
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(record_user=True,
                     record_request_args=True,
                     source='course_views[selectCourse]')
def selectCourse(request):
    """
    学生选课的聚合页面，包括: 
    1. 所有开放课程的选课信息
    2. 在预选和补退选阶段，学生可以通过点击课程对应的按钮实现选课或者退选，
    且点击后页面显示发生相应的变化
    3. 显示选课结果
    
    用户权限: 只有学生账号可以进入，组织和老师均不应该进入该页面
    """
    valid, user_type, html_display = utils.check_user_type(request.user)
    me = get_person_or_org(request.user, user_type)

    if (user_type == "Organization"
            or me.identity == NaturalPerson.Identity.TEACHER):
        return redirect(message_url(wrong("非学生账号不能选课！")))

    # 暂时不启用意愿点机制
    # if not is_staff:
    #     html_display["willing_point"] = remaining_willingness_point(me)

    # 学生选课或者取消选课

    if request.method == 'POST':

        # 参数: 课程id，操作action: select/cancel

        try:
            course_id = request.POST.get('courseid')
            action = request.POST.get('action')

            # 合法性检查
            assert action == "select" or action == "cancel"
            assert Course.objects.activated().filter(id=course_id).exists()

        except:
            wrong("出现预料之外的错误！如有需要，请联系管理员。", html_display)
        try:
            # 对学生的选课状态进行变更
            context = registration_status_change(course_id, me, action)
            my_messages.transfer_message_context(context, html_display)
        except:
            wrong("选课过程出现错误！请联系管理员。", html_display)

    html_display["is_myself"] = True
    html_display["current_year"] = CURRENT_ACADEMIC_YEAR
    html_display["semester"] = ("春" if Semester.now() == Semester.SPRING else "秋")

    html_display["yx_election_start"] = get_setting("course/yx_election_start")
    html_display["yx_election_end"] = get_setting("course/yx_election_end")
    html_display["btx_election_start"] = get_setting("course/btx_election_start")
    html_display["btx_election_end"] = get_setting("course/btx_election_end")

    # 是否正在进行抽签
    is_drawing = (str_to_time(html_display["yx_election_end"]) <= datetime.now()
                   <= str_to_time(html_display["btx_election_start"]))

    # 选课是否已经全部结束
    is_end = (datetime.now() > str_to_time(html_display["btx_election_end"]))

    unselected_courses = Course.objects.unselected(me)
    selected_courses = Course.objects.selected(me)

    # 未选的课程需要按照课程类型排序
    courses = {}
    for type, label in Course.CourseType.choices:
        # 前端使用键呈现
        courses[label] = course_to_display(unselected_courses.filter(type=type),
                                          me)

    unselected_display = course_to_display(unselected_courses, me)
    selected_display = course_to_display(selected_courses, me)

    bar_display = utils.get_sidebar_and_navbar(request.user, "书院课程")

    return render(request, "select_course.html", locals())


@login_required(redirect_field_name="origin")
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(record_user=True,
                     record_request_args=True,
                     source='course_views[viewCourse]')
def viewCourse(request):
    """
    展示一门课程的详细信息
    
    GET参数: ?courseid=<int>

    用户权限: 不对用户类型作出限制，均正常显示内容  
    """
    valid, user_type, html_display = utils.check_user_type(request.user)

    try:
        course_id = int(request.GET.get("courseid", None))
        course = Course.objects.filter(id=course_id)

        assert course.exists()

    except:
        return redirect(message_url(wrong("该课程不存在！")))

    me = utils.get_person_or_org(request.user, user_type)
    course_display = course_to_display(course, me, detail=True)

    bar_display = utils.get_sidebar_and_navbar(request.user, "课程详情")

    return render(request, "course_info.html", locals())


@login_required(redirect_field_name="origin")
@utils.check_user_access(redirect_url="/logout/")
@log.except_captured(EXCEPT_REDIRECT, source='course_views[addCourse]', record_user=True)
def addCourse(request, cid=None):
    """
    发起课程页
    ---------------
    页面逻辑：

    该函数处理 GET, POST 两种请求，发起和修改两类操作
    1. 访问 /addCourse/ 时，为创建操作，要求用户是小组；
    2. 访问 /editCourse/aid 时，为编辑操作，要求用户是该活动的发起者
    3. GET 请求创建课程的界面，placeholder 为 prompt
    4. GET 请求编辑课程的界面，表单的 placeholder 会被修改为课程的旧值。
    """

    # 检查：不是超级用户，必须是小组，修改是必须是自己
    try:
        valid, user_type, html_display = utils.check_user_type(request.user)
        # assert valid  已经在check_user_access检查过了
        me = utils.get_person_or_org(request.user, user_type) # 这里的me应该为小组账户
        if cid is None:
            if user_type != "Organization" or me.otype.otype_name != COURSE_TYPENAME:
                return redirect(message_url(wrong('书院课程账号才能发起课程!')))
            #暂时仅支持一个课程账号一学期只能开一门课
            courses = Course.objects.activated().filter(organization=me)
            if courses.exists():
                cid = courses[0].id
                return redirect(message_url(
                            succeed('您已在本学期创建过课程，已为您自动跳转!'),
                            f'/editCourse/{cid}'))
            edit = False
        else:
            cid = int(cid)
            course = Course.objects.get(id=cid)
            if course.organization != me:
                return redirect(message_url(wrong("无法修改其他小组的课程!")))
            edit = True
        html_display["is_myself"] = True
    except Exception as e:
        log.record_traceback(request, e)
        return EXCEPT_REDIRECT

    html_display["warn_code"] = int(request.GET.get("warn_code", 0))  # 是否有来自外部的消息
    html_display["warn_message"] = request.GET.get(
            "warn_message", "")  # 提醒的具体内容

    # 处理 POST 请求
    # 在这个界面，不会返回render，而是直接跳转到viewCourse，可以不设计bar_display
    if request.method == "POST" and request.POST:
        if not edit:

            #增加截止开课的时间点
            add_course_DDL = str_to_time(get_setting("course/btx_election_end"))
            if datetime.now() > add_course_DDL:
                return redirect(message_url(succeed("已超过选课时间节点，无法发起课程！"),
                                        f'/showCourseActivity/'))
            #发起选课
            context=create_course(request)
            html_display["warn_code"] = context["warn_code"]
            if html_display["warn_code"] == 2:
                return redirect(message_url(succeed("创建课程成功！为您自动跳转到编辑界面。"),
                                        f'/editCourse/{context["cid"]}'))
        else:
            # 仅未开始选课阶段可以修改
            if course.status != Course.Status.WAITING:
                return redirect(message_url(wrong('当前课程状态不允许修改!'),
                                            f'/editCourse/{course.id}'))
            context = create_course(request, course.id)
            course = Course.objects.get(id=context["cid"])
        html_display["warn_code"] = context["warn_code"]
        html_display["warn_message"] = context["warn_message"]

    # 下面的操作基本如无特殊说明，都是准备前端使用量
    html_display["applicant_name"] = me.oname
    html_display["app_avatar_path"] = me.get_user_ava()
    html_display["today"] = datetime.now().strftime("%Y-%m-%d")
    course_type_all = [
       ["德" , Course.CourseType.MORAL] ,
       ["智" , Course.CourseType.INTELLECTUAL] ,
       ["体" , Course.CourseType.PHYSICAL] ,
       ["美" , Course.CourseType.AESTHETICS],
       ["劳" , Course.CourseType.LABOUR],
    ]
    defaultpics = [{"src": f"/static/assets/img/announcepics/{i+1}.JPG", "id": f"picture{i+1}"} for i in range(5)]
    editable=False
    if edit and course.status == Course.Status.WAITING: #选课未开始才能修改
        editable = True

    if edit:
        name = utils.escape_for_templates(course.name)
        organization = course.organization
        year = course.year
        semester = utils.escape_for_templates(course.semester)
        classroom = utils.escape_for_templates(course.classroom)
        teacher = utils.escape_for_templates(course.teacher)
        course_time = course.time_set.all()
        introduction = utils.escape_for_templates(course.introduction)
        teaching_plan=utils.escape_for_templates(course.teaching_plan)
        record_cal_method=utils.escape_for_templates(course.record_cal_method)
        status = course.status
        capacity = course.capacity
        type = course.type
        current_participants = course.current_participants
        QRcode=course.QRcode


    if not edit:
        bar_display = utils.get_sidebar_and_navbar(request.user, "发起课程")
    else:
        bar_display = utils.get_sidebar_and_navbar(request.user, "修改课程")

    return render(request, "register_course.html", locals())


def downloadCourseRecord(me,year,semester):
    '''
    返回需要导出的文件
    '''
    year = CURRENT_ACADEMIC_YEAR
    semester = Semester.now()
    try:
        course = Course.objects.activated().get(organization = me)
    except:
        return redirect(message_url(wrong('未查询到相应课程，请联系管理员。')))

    records = CourseRecord.objects.filter(
        year = year,
        semester = semester,
        course = course,
    )
    if not records.exists():
        return redirect(message_url(wrong('未查询到相应课程记录，请联系管理员。')))

    wb = Workbook()		# 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active	# 获取第一个工作表（sheet1）
    sheet1.title = re.sub('[{}]'.format(punctuation),"",str(me.oname)) # 给工作表设置标题
    sheet_header = ['课程','姓名','学号','次数','学时',"学年","学期"]
    for i in range(len(sheet_header)):	# 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        sheet1.cell(row=1, column=i+1).value=sheet_header[i]
    max_row = sheet1.max_row
    for record in records:
        max_row += 1
        record_info = [
            str(me.oname),
            record.person.name, 
            str(record.person.person_id), 
            record.attend_times, 
            record.total_hours,
            str(year),
            str(semester) ]
        for x in range(len(record_info)):		# 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x+1).value = record_info[x]
            
    output = BytesIO()
    wb.save(output)	 # 将Excel文件内容保存到IO中
    output.seek(0)	 # 重新定位到开始
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = str(me.oname)+'-{}'.format(ctime)	# 给文件名中添加日期时间
    file_name = re.sub('[{}]'.format(punctuation),"",file_name) #去除中文符号
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = 'attachment;filename={}.xlsx'.format(file_name).encode('utf-8')
    wb.save(response)
    return response